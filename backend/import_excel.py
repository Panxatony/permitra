"""Importiert die bestehende Kommunikationsmatrix (Excel) in die Permitra-Datenbank.

Aufruf:
    python import_excel.py <datei.xlsx> [--sheet Kommunikationsmatrix] [--wipe]

Spalten-Mapping (wie AP0400-Sicherheitsregeln):
    Rule-ID | Application | Sicherheitselement | Source SZ | Quelle/Quellsystem |
    Destination-SZ | Ziel/Zielsystem | Protokoll | Port | Anlass | Requestor |
    Bearbeiter | Status Juniper | Status ACI | Status | Change-ID | Letzte Änderung |
    Info | Fachlicher Bezug
"""
import argparse
import sys

import openpyxl

import re

from app.database import Base, SessionLocal, engine
from app.models import ComponentType, Rule, RuleStatus, RuleVersion, SecurityComponent
from app.validation import extract_networks


def parse_address_lines(text: str) -> list[dict]:
    """Wandelt Alt-Format-Zeilen ("host.example.de - 10.40.72.5") in strukturierte
    Einträge {"ip": ..., "alias": ...} um. Zeilen ohne IP werden als Alias ohne IP
    übernommen, "any"/"Internet" wird zu ip="any"."""
    entries = []
    for line in (text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if line.lower() in ("any", "internet"):
            entries.append({"ip": "any", "alias": "" if line.lower() == "any" else line})
            continue
        host = re.split(r"[\s(]", line, maxsplit=1)[0].rstrip("-").strip()
        networks = extract_networks(line)
        if networks:
            alias = "" if host and extract_networks(host) else host
            for net in networks:
                ip = str(net.network_address) if net.num_addresses == 1 else str(net)
                entries.append({"ip": ip, "alias": alias})
        else:
            entries.append({"ip": "", "alias": line})
    return entries or [{"ip": "any", "alias": ""}]

IMPORT_COMPONENT_NAMES = {
    "juniper": "Juniper (Import)",
    "checkpoint": "Check Point (Import)",
    "aci": "ACI (Import)",
}


def component_for(db, cache: dict, platform: str) -> SecurityComponent:
    """Findet die erste Komponente des Typs oder legt eine Import-Komponente an."""
    if platform not in cache:
        component = (
            db.query(SecurityComponent)
            .filter(SecurityComponent.type == ComponentType(platform))
            .order_by(SecurityComponent.id)
            .first()
        )
        if not component:
            component = SecurityComponent(
                name=IMPORT_COMPONENT_NAMES[platform], type=ComponentType(platform),
                description="Automatisch beim Excel-Import angelegt",
            )
            db.add(component)
            db.flush()
        cache[platform] = component
    return cache[platform]

STATUS_MAP = {
    "umgesetzt": RuleStatus.approved,
    "neu": RuleStatus.in_review,
    "neui": RuleStatus.in_review,
    "deaktivieren": RuleStatus.deactivated,
    "deaktiviert": RuleStatus.deactivated,
}


def cell(row, idx) -> str:
    value = row[idx] if idx < len(row) else None
    return str(value).strip() if value is not None else ""


def parse_services(protocol_cell: str, port_cell: str) -> list[dict]:
    protos = [p.strip() for p in protocol_cell.splitlines() if p.strip()]
    ports = [p.strip() for p in port_cell.splitlines() if p.strip()]
    services = []

    if len(protos) == len(ports):
        pairs = list(zip(protos, ports))
    elif len(protos) == 1:
        pairs = [(protos[0], p) for p in ports] or [(protos[0], "")]
    elif len(ports) == 1:
        pairs = [(p, ports[0]) for p in protos]
    else:  # ungleich und mehrdeutig: bestmöglich paaren
        pairs = list(zip(protos, ports + [""] * (len(protos) - len(ports))))

    for proto, port in pairs:
        proto_norm = proto.upper().replace("ICMPV6", "ICMP").strip()
        port_norm = port.strip()
        if proto_norm.startswith("ICMP") or port_norm.lower().startswith(("icmp", "ping")):
            proto_norm, port_norm = "ICMP", ""
        if not proto_norm:
            proto_norm = "TCP"
        services.append({"protocol": proto_norm, "port": port_norm})
    return services or [{"protocol": "ANY", "port": "any"}]


def parse_platforms(element: str) -> list[str]:
    element = element.lower()
    platforms = []
    if "juniper" in element:
        platforms.append("juniper")
    if "check" in element:
        platforms.append("checkpoint")
    if "aci" in element:
        platforms.append("aci")
    return platforms


def run(path: str, sheet: str, wipe: bool):
    Base.metadata.create_all(bind=engine)
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"Sheet '{sheet}' nicht gefunden. Vorhanden: {wb.sheetnames}")
    ws = wb[sheet]

    db = SessionLocal()
    if wipe:
        db.query(RuleVersion).delete()
        db.query(Rule).delete()
        db.commit()

    imported, skipped, last_rule = 0, 0, None
    component_cache: dict = {}
    rows = ws.iter_rows(values_only=True)
    next(rows)  # Kopfzeile
    for row in rows:
        rule_id = cell(row, 0)
        if not rule_id:
            # Fortsetzungszeile: weiteres Ziel zur vorherigen Regel
            extra_dst = cell(row, 6)
            if last_rule is not None and extra_dst:
                last_rule.destination = list(last_rule.destination) + parse_address_lines(extra_dst)
            continue
        if db.query(Rule).filter(Rule.rule_id == rule_id).first():
            skipped += 1
            continue

        status = STATUS_MAP.get(cell(row, 14).lower(), RuleStatus.draft)
        rule_components = [
            component_for(db, component_cache, p) for p in parse_platforms(cell(row, 2))
        ]
        impl_status = {}
        if cell(row, 12):  # "Status Juniper"
            impl_status[component_for(db, component_cache, "juniper").name] = cell(row, 12).lower()
        if cell(row, 13):  # "Status ACI"
            impl_status[component_for(db, component_cache, "aci").name] = cell(row, 13).lower()

        rule = Rule(
            rule_id=rule_id,
            name=cell(row, 9)[:64] or rule_id,
            application=cell(row, 1),
            components=rule_components,
            source_zone=cell(row, 3),
            source=parse_address_lines(cell(row, 4)),
            destination_zone=cell(row, 5),
            destination=parse_address_lines(cell(row, 6)),
            services=parse_services(cell(row, 7), cell(row, 8)),
            justification=cell(row, 9),
            requestor=cell(row, 10),
            owner=cell(row, 11),
            status=status,
            impl_status=impl_status,
            change_id=cell(row, 15).replace("\n", ", "),
            info=cell(row, 17),
            business_context=cell(row, 18),
            created_by="excel-import",
        )
        db.add(rule)
        db.flush()
        db.add(
            RuleVersion(
                rule_pk=rule.id, version=1, snapshot={"import": "excel", "row": rule_id},
                change_note=f"Import aus {path.rsplit('/', 1)[-1]}", changed_by="excel-import",
            )
        )
        last_rule = rule
        imported += 1

    db.commit()
    db.close()
    print(f"Import fertig: {imported} Regeln importiert, {skipped} übersprungen (ID existiert).")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("file")
    parser.add_argument("--sheet", default="Kommunikationsmatrix")
    parser.add_argument("--wipe", action="store_true", help="Bestehende Regeln vorher löschen")
    args = parser.parse_args()
    run(args.file, args.sheet, args.wipe)
