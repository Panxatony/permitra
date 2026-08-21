"""Importiert die Zonen-Kommunikationsmatrix (Excel, Sheet "Netze intern und Zonen").

Aufruf:
    python import_zones.py <datei.xlsx> [--sheet "Netze intern und Zonen"] [--wipe]

Erwartetes Format: eine Kreuztabelle "Von SZ / Nach SZ" mit Zellwerten wie
    "Allow Only (FW)", "Allow Only (FW/ACI)", "Allow Only (FW) Temp",
    "Block All (FW)", "Block All (ACI)", "-" (Diagonale)
Kopfzeilen- und Zeilen-Schreibweisen werden normalisiert (T-VPNS == T-VPN-S).
"""
import argparse
import re
import sys

import openpyxl

from app.database import Base, SessionLocal, engine
from app.models import Zone, ZonePolicy, ZonePolicyType


def norm(name: str) -> str:
    """Vergleichsschlüssel: Groß, ohne Trennzeichen (T-VPN-S == T-VPNS)."""
    return re.sub(r"[^A-Z0-9]", "", (name or "").upper())


def parse_cell(text: str):
    """Liefert (policy, temporary) oder None für '-'/leer.

    Das Durchsetzungselement in Klammern ("(FW)", "(ACI)", "(FW/ACI)") wird bewusst
    ignoriert: zwischen Zonen steht immer eine Firewall, ACI ist nur intra-zonal.
    """
    text = (text or "").strip()
    if not text or text == "-":
        return None
    lower = text.lower()
    if lower.startswith("allow"):
        policy = ZonePolicyType.allow_only
    elif lower.startswith("block"):
        policy = ZonePolicyType.block_all
    else:
        return None
    return policy, "temp" in lower


def run(path: str, sheet_name: str | None, wipe: bool):
    Base.metadata.create_all(bind=engine)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Kopfzeile finden ("Von SZ / Nach SZ")
    header_idx = next(
        (i for i, r in enumerate(rows) if r and str(r[0] or "").strip().lower().startswith("von sz")),
        None,
    )
    if header_idx is None:
        sys.exit("Kopfzeile 'Von SZ / Nach SZ' nicht gefunden")
    header = rows[header_idx]
    col_names = [str(c).strip() for c in header[1:] if c and str(c).strip()]

    db = SessionLocal()
    if wipe:
        db.query(ZonePolicy).delete()
        db.query(Zone).delete()
        db.commit()

    # Zonen anlegen: Zeilen-Schreibweise ist führend, Spalten ergänzen fehlende
    zones_by_key: dict[str, Zone] = {
        norm(z.name): z for z in db.query(Zone).all()
    }
    order = 0
    row_names = [str(r[0]).strip() for r in rows[header_idx + 1:] if r and r[0]]
    for name in row_names + col_names:
        key = norm(name)
        if key and key not in zones_by_key:
            zone = Zone(name=name, sort_order=order)
            db.add(zone)
            db.flush()
            zones_by_key[key] = zone
        order += 1

    # Matrix einlesen
    count = 0
    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        from_zone = zones_by_key.get(norm(str(row[0])))
        if not from_zone:
            continue
        for col_idx, col_name in enumerate(col_names, start=1):
            to_zone = zones_by_key.get(norm(col_name))
            if not to_zone or to_zone.id == from_zone.id:
                continue
            parsed = parse_cell(str(row[col_idx]) if col_idx < len(row) and row[col_idx] else "")
            if not parsed:
                continue
            policy_type, temporary = parsed
            existing = (
                db.query(ZonePolicy)
                .filter(ZonePolicy.from_zone_id == from_zone.id, ZonePolicy.to_zone_id == to_zone.id)
                .first()
            )
            if not existing:
                existing = ZonePolicy(from_zone_id=from_zone.id, to_zone_id=to_zone.id)
                db.add(existing)
            existing.policy = policy_type
            existing.temporary = temporary
            count += 1

    db.commit()
    zones_total = db.query(Zone).count()
    db.close()
    print(f"Import fertig: {zones_total} Zonen, {count} Zonen-Beziehungen gepflegt.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("file")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--wipe", action="store_true", help="Zonen + Matrix vorher löschen")
    args = parser.parse_args()
    run(args.file, args.sheet, args.wipe)
